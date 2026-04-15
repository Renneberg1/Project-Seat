"""Tests for the Documentation Run Sheet Excel export service."""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from src.models.dhf import DHFDocument, DocumentStatus
from src.services.runsheet_export import (
    PHASE_AREA_TO_NUMBER,
    RunsheetExportService,
    get_phase_number,
    strip_area_prefix,
)


def _make_doc(
    title: str = "Test Doc",
    area: str = "Software",
    released_version: str | None = "1.0",
    draft_version: str | None = None,
    status: DocumentStatus = DocumentStatus.RELEASED,
    page_url: str = "https://example.com/draft/page",
    released_page_url: str | None = "https://example.com/released/page",
) -> DHFDocument:
    return DHFDocument(
        title=title,
        area=area,
        released_version=released_version,
        draft_version=draft_version,
        status=status,
        last_modified="2026-01-15T10:00:00Z",
        author="Test Author",
        page_url=page_url,
        released_page_url=released_page_url,
    )


@pytest.fixture
def service() -> RunsheetExportService:
    return RunsheetExportService()


class TestStripAreaPrefix:
    def test_strips_radx_prefix(self) -> None:
        assert strip_area_prefix("RadX Software") == "Software"

    def test_strips_product_prefix(self) -> None:
        assert strip_area_prefix("RadX Verification & Validation") == "Verification & Validation"

    def test_no_prefix(self) -> None:
        assert strip_area_prefix("Software") == "Software"

    def test_unknown_remainder_kept(self) -> None:
        assert strip_area_prefix("RadX Foobar") == "RadX Foobar"

    def test_strips_design_inputs(self) -> None:
        assert strip_area_prefix("RadX Design Inputs") == "Design Inputs"


class TestGetPhaseNumber:
    def test_known_areas(self) -> None:
        assert get_phase_number("Planning") == 1
        assert get_phase_number("Design Input") == 2
        assert get_phase_number("Design Inputs") == 2
        assert get_phase_number("Software") == 3
        assert get_phase_number("Risk Management") == 4
        assert get_phase_number("Verification and Validation") == 5
        assert get_phase_number("Guides") == 6
        assert get_phase_number("Design Transfer") == 7
        assert get_phase_number("Quality Assurance") == 8

    def test_case_insensitive(self) -> None:
        assert get_phase_number("planning") == 1
        assert get_phase_number("PLANNING") == 1

    def test_unknown_area_returns_fallback(self) -> None:
        assert get_phase_number("Unknown Area") == 99


class TestRunsheetExportService:
    def test_generates_valid_xlsx(self, service: RunsheetExportService) -> None:
        docs = [_make_doc()]
        buf = service.generate("Project X", docs, {"Test Doc": "To Do"}, "v1.0")
        wb = load_workbook(buf)
        assert wb.sheetnames == ["Revision History", "Runsheet"]

    def test_runsheet_header_row(self, service: RunsheetExportService) -> None:
        buf = service.generate("P", [_make_doc()], {"Test Doc": "Published"}, "v1")
        wb = load_workbook(buf)
        ws = wb["Runsheet"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
        assert headers == ["Phase", "Phase Area", "Document", "Version", "Status"]

    def test_runsheet_has_table(self, service: RunsheetExportService) -> None:
        buf = service.generate("P", [_make_doc()], {"Test Doc": "To Do"}, "v1")
        wb = load_workbook(buf)
        ws = wb["Runsheet"]
        assert len(ws.tables) == 1
        for tbl in ws.tables.values():
            assert "A1:" in tbl.ref
            assert tbl.tableStyleInfo.name == "TableStyleMedium16"
            assert tbl.tableStyleInfo.showRowStripes is True

    def test_runsheet_has_conditional_formatting(self, service: RunsheetExportService) -> None:
        docs = [_make_doc(title="A"), _make_doc(title="B")]
        buf = service.generate("P", docs, {"A": "Published", "B": "No Change"}, "v1")
        wb = load_workbook(buf)
        ws = wb["Runsheet"]
        # Should have 4 conditional formatting rules (Published, To Do, Pending, No Change)
        rule_count = sum(len(cf.rules) for cf in ws.conditional_formatting)
        assert rule_count == 4

    def test_document_count_matches_input(self, service: RunsheetExportService) -> None:
        docs = [
            _make_doc(title="Doc A", area="Software"),
            _make_doc(title="Doc B", area="Clinical"),
            _make_doc(title="Doc C", area="Planning"),
        ]
        status_map = {d.title: "To Do" for d in docs}
        buf = service.generate("P", docs, status_map, "v1")
        wb = load_workbook(buf)
        ws = wb["Runsheet"]
        data_rows = [ws.cell(row=r, column=3).value for r in range(2, 100) if ws.cell(row=r, column=3).value]
        assert len(data_rows) == 3

    def test_phase_number_from_area(self, service: RunsheetExportService) -> None:
        docs = [_make_doc(title="SRS", area="Software")]
        buf = service.generate("P", docs, {"SRS": "To Do"}, "v1")
        wb = load_workbook(buf)
        ws = wb["Runsheet"]
        assert ws.cell(row=2, column=1).value == 3
        assert ws.cell(row=2, column=2).value == "Software"

    def test_design_inputs_maps_to_phase_2(self, service: RunsheetExportService) -> None:
        docs = [_make_doc(title="Spec", area="RadX Design Inputs")]
        buf = service.generate("P", docs, {"Spec": "To Do"}, "v1")
        wb = load_workbook(buf)
        ws = wb["Runsheet"]
        assert ws.cell(row=2, column=1).value == 2
        assert ws.cell(row=2, column=2).value == "Design Inputs"

    def test_phase_area_strips_prefix(self, service: RunsheetExportService) -> None:
        docs = [_make_doc(title="SRS", area="RadX Software")]
        buf = service.generate("P", docs, {"SRS": "To Do"}, "v1")
        wb = load_workbook(buf)
        ws = wb["Runsheet"]
        assert ws.cell(row=2, column=2).value == "Software"
        assert ws.cell(row=2, column=1).value == 3

    def test_unmapped_area_gets_fallback(self, service: RunsheetExportService) -> None:
        docs = [_make_doc(title="Mystery", area="Unknown Area")]
        buf = service.generate("P", docs, {"Mystery": "No Change"}, "v1")
        wb = load_workbook(buf)
        ws = wb["Runsheet"]
        assert ws.cell(row=2, column=1).value == 99
        assert ws.cell(row=2, column=2).value == "Unknown Area"

    def test_status_values_written(self, service: RunsheetExportService) -> None:
        docs = [
            _make_doc(title="A"),
            _make_doc(title="B"),
            _make_doc(title="C"),
            _make_doc(title="D"),
        ]
        status_map = {
            "A": "Published",
            "B": "Pending",
            "C": "To Do",
            "D": "No Change",
        }
        buf = service.generate("P", docs, status_map, "v1")
        wb = load_workbook(buf)
        ws = wb["Runsheet"]
        statuses = {
            ws.cell(row=r, column=3).value: ws.cell(row=r, column=5).value
            for r in range(2, 6)
        }
        assert statuses["A"] == "Published"
        assert statuses["B"] == "Pending"
        assert statuses["C"] == "To Do"
        assert statuses["D"] == "No Change"

    def test_revision_history_layout(self, service: RunsheetExportService) -> None:
        buf = service.generate("My Project", [], {}, "v2.0")
        wb = load_workbook(buf)
        ws = wb["Revision History"]
        assert ws["A1"].value == "Version"
        assert ws["B1"].value == "Description of Change"
        assert ws["C1"].value == "Updated By"
        assert ws["D1"].value == "Date"
        assert ws["A2"].value == 1
        assert "Initial Assessment" in ws["B2"].value
        assert ws["C2"].value == "My Project"

    def test_revision_history_purpose(self, service: RunsheetExportService) -> None:
        buf = service.generate("My Project", [], {}, "v2.0")
        wb = load_workbook(buf)
        ws = wb["Revision History"]
        assert ws["A6"].value == "Purpose"
        purpose = ws["B6"].value
        assert "tracks the documentation deliverables" in purpose
        assert "My Project" in purpose
        assert "all activities have been completed prior to release" in purpose

    def test_revision_history_abbreviations(self, service: RunsheetExportService) -> None:
        buf = service.generate("P", [], {}, "v1")
        wb = load_workbook(buf)
        ws = wb["Revision History"]
        assert ws["A8"].value == "Abbreviation"
        assert ws["B8"].value == "Document Category"
        assert ws["A9"].value == "HAI"
        assert ws["B9"].value == "Harrison"

    def test_sorted_by_phase_area_title(self, service: RunsheetExportService) -> None:
        docs = [
            _make_doc(title="Zebra", area="Clinical"),
            _make_doc(title="Alpha", area="Software"),
            _make_doc(title="Beta", area="Planning"),
        ]
        status_map = {d.title: "To Do" for d in docs}
        buf = service.generate("P", docs, status_map, "v1")
        wb = load_workbook(buf)
        ws = wb["Runsheet"]
        titles = [ws.cell(row=r, column=3).value for r in range(2, 5)]
        assert titles[0] == "Beta"  # phase 1

    def test_version_falls_back_to_draft(self, service: RunsheetExportService) -> None:
        doc = _make_doc(title="Draft Only", released_version=None, draft_version="0.3")
        buf = service.generate("P", [doc], {"Draft Only": "To Do"}, "v1")
        wb = load_workbook(buf)
        ws = wb["Runsheet"]
        assert ws.cell(row=2, column=4).value == "0.3"

    def test_empty_documents(self, service: RunsheetExportService) -> None:
        buf = service.generate("P", [], {}, "v1")
        wb = load_workbook(buf)
        assert "Runsheet" in wb.sheetnames

    def test_document_hyperlink_to_released_url(self, service: RunsheetExportService) -> None:
        doc = _make_doc(
            title="SRS",
            released_page_url="https://confluence.example.com/released/123",
        )
        buf = service.generate("P", [doc], {"SRS": "Published"}, "v1")
        wb = load_workbook(buf)
        ws = wb["Runsheet"]
        cell = ws.cell(row=2, column=3)
        assert cell.hyperlink is not None
        assert cell.hyperlink.target == "https://confluence.example.com/released/123"

    def test_document_hyperlink_falls_back_to_page_url(self, service: RunsheetExportService) -> None:
        doc = _make_doc(
            title="Draft Doc",
            released_page_url=None,
            page_url="https://confluence.example.com/draft/456",
        )
        buf = service.generate("P", [doc], {"Draft Doc": "To Do"}, "v1")
        wb = load_workbook(buf)
        ws = wb["Runsheet"]
        cell = ws.cell(row=2, column=3)
        assert cell.hyperlink is not None
        assert cell.hyperlink.target == "https://confluence.example.com/draft/456"
