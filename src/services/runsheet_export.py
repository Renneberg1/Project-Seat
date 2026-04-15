"""Documentation Run Sheet Excel export service."""

from __future__ import annotations

import re
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.models.dhf import DHFDocument

# Phase area name (lowercase) → phase number.
# Phase area = DHF area with product prefix stripped (e.g. "RadX Software" → "Software").
PHASE_AREA_TO_NUMBER: dict[str, int] = {
    # Phase 1 — Planning
    "planning": 1,
    "project management": 1,
    # Phase 2 — Design Input(s)
    "design input": 2,
    "design inputs": 2,
    "product management": 2,
    "regulatory affairs": 2,
    # Phase 3 — Design Outputs
    "design outputs": 3,
    "system": 3,
    "software": 3,
    "deployment": 3,
    "clinical": 3,
    "clinical research and evaluation": 3,
    # Phase 4 — Risk and Usability
    "risk and usability": 4,
    "risk management": 4,
    # Phase 5 — Verification and Validation
    "verification and validation": 5,
    "verification & validation": 5,
    "post market surveillance": 5,
    # Phase 6 — Guides and Release Notes
    "guides": 6,
    "guides and release notes": 6,
    # Phase 7 — Design Transfer
    "design transfer": 7,
    "design transfer and release": 7,
    # Phase 8 — Regulatory and Quality
    "regulatory and quality": 8,
    "raqa compliance": 8,
    "quality assurance": 8,
}

_FALLBACK_PHASE = 99

# ── Revision History styles (blue header fill matching example) ──────────
_FILL_HEADER = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
_ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_CENTER = Alignment(horizontal="center")
_ALIGN_LEFT = Alignment(horizontal="left")

_FONT_HEADER = Font(name="Calibri", bold=True, size=11)
_FONT_NORMAL = Font(name="Calibri", size=11)
_FONT_LINK = Font(name="Calibri", size=11, color="0563C1", underline="single")

# ── Conditional formatting fills for the Status column ───────────────────
_CF_PUBLISHED = PatternFill(bgColor="A9D18E")       # green
_CF_TODO = PatternFill(bgColor="FCE4D6")             # light orange
_CF_PENDING = PatternFill(bgColor="FFFFEB9C")        # amber
_CF_NO_CHANGE = PatternFill(bgColor="D9D9D9")        # gray
_CF_NO_CHANGE_FONT = Font(color="808080")

# Abbreviation / Document Category legend (matches example template)
ABBREVIATIONS = [
    ("HAI", "Harrison"),
    ("PRJ", "Project Management"),
    ("PRM", "Product Management"),
    ("SYS", "System"),
    ("SW", "Software"),
    ("DPL", "Deployment"),
    ("CLI", "Clinical"),
    ("RA", "Regulatory Affairs"),
    ("QA", "Quality Assurance"),
    ("VNV", "Verification & Validation"),
]

# Regex to strip a product-name prefix from area names
_PRODUCT_PREFIX_RE = re.compile(r"^[A-Z][a-zA-Z0-9]*[-\s]+", re.ASCII)

PURPOSE_TEXT = (
    "This record tracks the documentation deliverables for {project_name}. "
    "Prior to commencing development, as part of the release planning all "
    "documents requiring update have been identified. This record is then "
    "tracked the updates, reviews and approvals of these document to ensure "
    "all activities have been completed prior to release."
)


def strip_area_prefix(area: str) -> str:
    """Strip a leading product-name prefix from a DHF area name.

    Examples:
        "RadX Software" → "Software"
        "RadX Verification & Validation" → "Verification & Validation"
        "Software" → "Software"  (no prefix to strip)
    """
    m = _PRODUCT_PREFIX_RE.match(area)
    if m:
        remainder = area[m.end():]
        if remainder.lower() in PHASE_AREA_TO_NUMBER:
            return remainder
    return area


def get_phase_number(phase_area: str) -> int:
    """Look up the phase number for a cleaned area name."""
    return PHASE_AREA_TO_NUMBER.get(phase_area.lower(), _FALLBACK_PHASE)


class RunsheetExportService:
    """Generate a Documentation Run Sheet as an Excel workbook."""

    def generate(
        self,
        project_name: str,
        documents: list[DHFDocument],
        status_map: dict[str, str],
        release_name: str = "",
    ) -> BytesIO:
        """Build an Excel workbook with Revision History and Runsheet sheets.

        Args:
            project_name: Display name for the project.
            documents: Full DHF document list.
            status_map: Mapping of doc title → status string
                        ("Published", "Pending", "To Do", "No Change").
            release_name: Name of the release (for the revision history).

        Returns:
            BytesIO containing the .xlsx file.
        """
        wb = Workbook()

        # --- Revision History sheet ---
        ws_rev = wb.active
        ws_rev.title = "Revision History"
        self._build_revision_history(ws_rev, project_name, release_name)

        # --- Runsheet sheet ---
        ws_run = wb.create_sheet("Runsheet")
        self._build_runsheet(ws_run, documents, status_map)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ------------------------------------------------------------------
    # Revision History sheet
    # ------------------------------------------------------------------

    def _build_revision_history(
        self, ws, project_name: str, release_name: str,
    ) -> None:
        # Row 1: Revision table header
        rev_headers = ["Version", "Description of Change", "Updated By", "Date"]
        for col_idx, header in enumerate(rev_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = _FONT_HEADER
            cell.fill = _FILL_HEADER
            cell.alignment = _ALIGN_CENTER_WRAP

        # Row 2: first revision entry
        r2_data = [
            1,
            "Initial Assessment for documentation deliverables",
            project_name,
            date.today().strftime("%B %Y"),
        ]
        for col_idx, val in enumerate(r2_data, 1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.font = _FONT_NORMAL
            cell.alignment = _ALIGN_CENTER_WRAP

        # Row 6: Purpose
        purpose_label = ws.cell(row=6, column=1, value="Purpose")
        purpose_label.font = _FONT_HEADER
        purpose_label.fill = _FILL_HEADER
        purpose_label.alignment = _ALIGN_CENTER_WRAP

        purpose_cell = ws.cell(
            row=6, column=2,
            value=PURPOSE_TEXT.format(project_name=project_name),
        )
        purpose_cell.font = _FONT_NORMAL
        purpose_cell.alignment = _ALIGN_LEFT_WRAP
        ws.merge_cells("B6:D6")

        # Row 8: Abbreviation / Document Category table
        abbr_headers = ["Abbreviation", "Document Category"]
        for col_idx, header in enumerate(abbr_headers, 1):
            cell = ws.cell(row=8, column=col_idx, value=header)
            cell.font = _FONT_HEADER
            cell.fill = _FILL_HEADER
            cell.alignment = _ALIGN_CENTER_WRAP

        for row_idx, (abbr, defn) in enumerate(ABBREVIATIONS, 9):
            cell_a = ws.cell(row=row_idx, column=1, value=abbr)
            cell_a.font = _FONT_NORMAL
            cell_a.alignment = _ALIGN_CENTER_WRAP
            cell_b = ws.cell(row=row_idx, column=2, value=defn)
            cell_b.font = _FONT_NORMAL
            cell_b.alignment = _ALIGN_LEFT_WRAP

        # Column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 15

    # ------------------------------------------------------------------
    # Runsheet sheet — Excel Table + conditional formatting
    # ------------------------------------------------------------------

    def _build_runsheet(
        self,
        ws,
        documents: list[DHFDocument],
        status_map: dict[str, str],
    ) -> None:
        headers = ["Phase", "Phase Area", "Document", "Version", "Status"]
        col_widths = [10, 29, 67, 10, 18]

        # Header row
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # Set column widths
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Build rows sorted by phase → phase area → title
        rows = []
        url_map: dict[str, str] = {}
        for doc in documents:
            phase_area = strip_area_prefix(doc.area)
            phase_num = get_phase_number(phase_area)
            status = status_map.get(doc.title, "No Change")
            version = doc.released_version or doc.draft_version or "—"
            rows.append((phase_num, phase_area, doc.title, version, status))
            if doc.released_page_url:
                url_map[doc.title] = doc.released_page_url
            elif doc.page_url:
                url_map[doc.title] = doc.page_url

        rows.sort(key=lambda r: (r[0], r[1], r[2]))

        # Write data rows
        for row_idx, (phase_num, phase_area, title, version, status) in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=phase_num)
            ws.cell(row=row_idx, column=2, value=phase_area)
            doc_cell = ws.cell(row=row_idx, column=3, value=title)
            doc_cell.alignment = _ALIGN_LEFT
            # Hyperlink to released/published Confluence page
            url = url_map.get(title)
            if url:
                doc_cell.hyperlink = url
                doc_cell.font = _FONT_LINK
            ws.cell(row=row_idx, column=4, value=version).alignment = _ALIGN_CENTER
            ws.cell(row=row_idx, column=5, value=status).alignment = _ALIGN_CENTER

        last_row = max(len(rows) + 1, 2)  # at least row 2 for table ref

        # Apply Excel Table with blue pivot-table style (TableStyleMedium16)
        table_ref = f"A1:E{last_row}"
        table = Table(displayName="Runsheet", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium16",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

        # Conditional formatting on the Status column (E)
        status_range = f"E2:E{last_row}"

        ws.conditional_formatting.add(
            status_range,
            Rule(
                type="containsText",
                operator="containsText",
                text="Published",
                formula=[f'NOT(ISERROR(SEARCH("Published",E2)))'],
                dxf=_dxf(fill=_CF_PUBLISHED),
            ),
        )
        ws.conditional_formatting.add(
            status_range,
            Rule(
                type="containsText",
                operator="containsText",
                text="To Do",
                formula=[f'NOT(ISERROR(SEARCH("To Do",E2)))'],
                dxf=_dxf(fill=_CF_TODO),
            ),
        )
        ws.conditional_formatting.add(
            status_range,
            Rule(
                type="containsText",
                operator="containsText",
                text="Pending",
                formula=[f'NOT(ISERROR(SEARCH("Pending",E2)))'],
                dxf=_dxf(fill=_CF_PENDING),
            ),
        )
        ws.conditional_formatting.add(
            status_range,
            Rule(
                type="containsText",
                operator="containsText",
                text="No Change",
                formula=[f'NOT(ISERROR(SEARCH("No Change",E2)))'],
                dxf=_dxf(fill=_CF_NO_CHANGE, font=_CF_NO_CHANGE_FONT),
            ),
        )


def _dxf(
    fill: PatternFill | None = None,
    font: Font | None = None,
) -> "DifferentialStyle":
    """Build an openpyxl DifferentialStyle for conditional formatting."""
    from openpyxl.styles.differential import DifferentialStyle

    return DifferentialStyle(fill=fill, font=font)
